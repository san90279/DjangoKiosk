from django.shortcuts import render
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.views.decorators.csrf import csrf_protect
import os
import datetime,json
from django.http import HttpResponse
from ExportExcel.models import M_V_PenaltyReport,M_V_FeeItemReport,SetDayReportExcel,SetMonthReportExcel
from Deal.models import M_DealMaster
from FeeItem.models import M_FeeItem
from CommonApp.models import GridCS
from django.core.serializers.json import DjangoJSONEncoder
from django.contrib import messages

# Create your views here.
def V_DayReportIndex(request):
    if(request.method=='GET'):
        return render(request,'ExportExcel/DayReportIndex.html');
    module_dir = os.path.dirname(__file__)
    path=open(os.path.join(module_dir+'/ExcelTemplate/', '日報表.xlsx'),'rb')
    wb = load_workbook(filename = path)
    ws = wb.active

    ws=SetDayReportExcel(request.POST.get('ExportDate', ''),'',ws)
    if(ws):
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename={0}.xlsx'.format(request.POST.get('ExportDate', ''))
        return response
    else:
        messages.success(request, '查無資料!', extra_tags='alert')
        return render(request,'ExportExcel/DayReportIndex.html');

def V_IntervalReportIndex(request):
    if(request.method=='GET'):
        return render(request,'ExportExcel/IntervalReportIndex.html');
    module_dir = os.path.dirname(__file__)
    path=open(os.path.join(module_dir+'/ExcelTemplate/', '日報表.xlsx'),'rb')
    wb = load_workbook(filename = path)
    ws = wb.active

    ws=SetDayReportExcel(request.POST.get('ExportDateS', ''),request.POST.get('ExportDateE', ''),ws)
    if(ws):
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename={0}_{1}.xlsx'.format(request.POST.get('ExportDateS', ''),request.POST.get('ExportDateE', ''))
        return response
    else:
        messages.success(request, '查無資料!', extra_tags='alert')
        return render(request,'ExportExcel/IntervalReportIndex.html');


def V_MonthReportIndex(request):
    MonthList=()
    YearList=()
    i=1
    m=1
    y=datetime.datetime.today().year
    while m<=12:
        MonthList=MonthList+((str(m),m),)
        m+=1
    while i<=5:
        YearList=YearList+((str(y),y),)
        y-=1
        i+=1
    if(request.method=='GET'):
        return render(request,'ExportExcel/MonthReportIndex.html',{"MonthList":MonthList,"YearList":YearList});
    module_dir = os.path.dirname(__file__)
    path=open(os.path.join(module_dir+'/ExcelTemplate/', '月報表.xlsx'),'rb')
    wb = load_workbook(filename = path)
    ws = wb.active

    ws=SetMonthReportExcel(request.POST.get('SelectYear', ''),request.POST.get('SelectMonth', ''),ws)
    if(ws):
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename={0}.xlsx'.format('Export')
        return response
    else:
        messages.success(request, '查無資料!', extra_tags='alert')
        return render(request,'ExportExcel/MonthReportIndex.html',{"MonthList":MonthList,"YearList":YearList});


#罰緩報表主頁
def V_PenaltyReportIndex(request):
    return render(request,'ExportExcel/PenaltyReportIndex.html');

#規費報表主頁
def V_FeeItemReportIndex(request):
    Status=M_DealMaster.DealStatusList
    PayType=M_DealMaster.PayTypeList
    Fee=M_FeeItem.objects.all().values_list('FeeID', 'FeeName')
    return render(request,'ExportExcel/FeeItemReportIndex.html',{'Status':Status,'PayType':PayType,'Fee':Fee});

#JQGRID取得罰緩報表資料
@csrf_protect
def V_GetPenaltyReport(request):
    draw = int(request.POST.get('draw'))  # 記錄操作次數
    #將前端request物件傳入GridCS內做處理
    grid=GridCS(request)
    #將Model M_V_PenaltyReport傳入作查詢
    PenaltyData=grid.dynamic_query_order(M_V_PenaltyReport)
    #將FeeItemData傳入作分頁
    object_list = grid.dynamic_query_order_paginator(PenaltyData)
    #資料總筆數
    count=len(PenaltyData)
    #拼出teplate JQGRID 欄位JSON資料流
    data=[{	'DealDate': Penalty.DealDate.strftime('%Y-%m-%d'),
            'PenaltyID':Penalty.PenaltyID  ,
            'PenaltyName': Penalty.PenaltyName,
            'TermID': Penalty.TermID ,
            'TermName': Penalty.TermName ,
            'Qty': Penalty.qty,
            'TotalAmount': Penalty.totalamount,
            'Remark': Penalty.Remark} for Penalty in object_list]
    #JQGRID API
    dic = {
        'draw': draw,
        'recordsTotal': count,
        'recordsFiltered': count,
        'data': data,
    }
    return HttpResponse(json.dumps(dic, cls=DjangoJSONEncoder), content_type='application/json')

#JQGRID取得規費報表資料
@csrf_protect
def V_GetFeeItemReport(request):
    draw = int(request.POST.get('draw'))  # 記錄操作次數
    #將前端request物件傳入GridCS內做處理
    grid=GridCS(request)
    #將Model M_V_FeeItemReport傳入作查詢
    FeeItemData=grid.dynamic_query_order(M_V_FeeItemReport)
    #將FeeItemData傳入作分頁
    object_list = grid.dynamic_query_order_paginator(FeeItemData)
    #資料總筆數
    count=len(FeeItemData)
    #拼出teplate JQGRID 欄位JSON資料流
    data=[{	'StationID': FeeItem.StationID,
            'DealDate': FeeItem.DealDate.strftime('%Y-%m-%d'),
            'DealTime': FeeItem.DealDate.strftime('%H:%M'),
            'InvoiceNo':FeeItem.InvoiceNo ,
            'Status': [val for key,val in M_DealMaster.DealStatusList if key==FeeItem.Status],
            'PayType':[val for key,val in M_DealMaster.PayTypeList if key==FeeItem.PayType],
            'FeeID': FeeItem.FeeID,
            'FeeName': FeeItem.FeeName,
            'Amount': FeeItem.Amount,
            'cAmount': FeeItem.Amount,
            'Qty': FeeItem.Qty} for FeeItem in object_list]
    #JQGRID API
    dic = {
        'draw': draw,
        'recordsTotal': count,
        'recordsFiltered': count,
        'data': data,
    }
    return HttpResponse(json.dumps(dic, cls=DjangoJSONEncoder), content_type='application/json')
